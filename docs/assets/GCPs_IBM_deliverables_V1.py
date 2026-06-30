"""
GCP Processor Script
====================
Converts GCP/PGCP files (Excel or CSV) into:
  - Shapefile in GCS WGS84
  - Shapefile in PCS (UTM, user-specified EPSG)
  - Excel with all columns including Lat/Lon, DMS, and Remarks

Usage:
    python gcp_processor.py --input <file.xlsx or file.csv>
                            --utm_epsg <EPSG code, e.g. 32644>
                            --mine_name <MineName>
                            --output_dir <output folder>

EPSG codes for common UTM zones (WGS84):
  UTM Zone 43N → 32643
  UTM Zone 44N → 32644
  UTM Zone 45N → 32645
  UTM Zone 46N → 32646

GCP Classification Logic:
  - Name starts with "GCP" (case-insensitive) → Temporary GCP
  - Any other name                             → Permanent GCP
"""

import os
import sys
import math
import pandas as pd
import shapefile  # pyshp
from pyproj import Transformer


# ─────────────────────────────────────────────
# 1. COORDINATE CONVERSION UTILITIES
# ─────────────────────────────────────────────

def utm_to_latlon(easting, northing, utm_epsg):
    """Convert UTM (Easting, Northing) to WGS84 (Lon, Lat)."""
    transformer = Transformer.from_crs(f"EPSG:{utm_epsg}", "EPSG:4326", always_xy=True)
    lon, lat = transformer.transform(easting, northing)
    return lat, lon


def dd_to_dms(decimal_degrees):
    """Convert decimal degrees to (degrees, minutes, seconds, direction_char)."""
    is_negative = decimal_degrees < 0
    decimal_degrees = abs(decimal_degrees)
    degrees = int(decimal_degrees)
    minutes_float = (decimal_degrees - degrees) * 60
    minutes = int(minutes_float)
    seconds = (minutes_float - minutes) * 60
    return degrees, minutes, seconds, is_negative


def format_dms_lat(lat):
    """Format latitude as DMS string e.g. 24°15'30.25\"N"""
    d, m, s, is_neg = dd_to_dms(lat)
    direction = "S" if is_neg else "N"
    return f"{d}°{m:02d}'{s:06.3f}\"{direction}"


def format_dms_lon(lon):
    """Format longitude as DMS string e.g. 082°30'15.10\"E"""
    d, m, s, is_neg = dd_to_dms(lon)
    direction = "W" if is_neg else "E"
    return f"{d}°{m:02d}'{s:06.3f}\"{direction}"


# ─────────────────────────────────────────────
# 2. REMARK CLASSIFICATION
# ─────────────────────────────────────────────

def classify_gcp(name):
    """Return remark based on GCP name."""
    if str(name).strip().upper().startswith("GCP"):
        return "Temporary GCP"
    else:
        return "Permanent GCP"


# ─────────────────────────────────────────────
# 3. READ INPUT FILE
# ─────────────────────────────────────────────

def read_input(file_path):
    """Read Excel or CSV into a DataFrame."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext in [".xlsx", ".xls"]:
        df = pd.read_excel(file_path)
    elif ext == ".csv":
        df = pd.read_csv(file_path)
    else:
        raise ValueError(f"Unsupported file format: {ext}. Use .xlsx, .xls, or .csv")

    # Normalize column names: strip whitespace
    df.columns = df.columns.str.strip()

    # Detect key columns (case-insensitive)
    col_map = {}
    for col in df.columns:
        cl = col.lower()
        if cl == "name":
            col_map["name"] = col
        elif cl in ["easting", "east", "e"]:
            col_map["easting"] = col
        elif cl in ["northing", "north", "n"]:
            col_map["northing"] = col
        elif cl in ["elevation", "elev", "height", "z", "altitude"]:
            col_map["elevation"] = col

    missing = [k for k in ["name", "easting", "northing"] if k not in col_map]
    if missing:
        raise ValueError(
            f"Could not find required columns: {missing}. "
            f"Columns found: {list(df.columns)}"
        )

    return df, col_map


# ─────────────────────────────────────────────
# 4. PROCESS DATA
# ─────────────────────────────────────────────

def process_data(df, col_map, utm_epsg):
    """Add Latitude, Longitude, DMS columns, and Remarks to DataFrame."""
    name_col = col_map["name"]
    east_col = col_map["easting"]
    north_col = col_map["northing"]

    lats, lons, dms_lats, dms_lons, remarks = [], [], [], [], []

    for _, row in df.iterrows():
        easting = float(row[east_col])
        northing = float(row[north_col])
        lat, lon = utm_to_latlon(easting, northing, utm_epsg)

        lats.append(round(lat, 8))
        lons.append(round(lon, 8))
        dms_lats.append(format_dms_lat(lat))
        dms_lons.append(format_dms_lon(lon))
        remarks.append(classify_gcp(row[name_col]))

    df = df.copy()
    df["Latitude"] = lats
    df["Longitude"] = lons
    df["DMS_Latitude"] = dms_lats
    df["DMS_Longitude"] = dms_lons
    df["Remarks"] = remarks

    return df


# ─────────────────────────────────────────────
# 5. WRITE SHAPEFILE (GCS WGS84)
# ─────────────────────────────────────────────

def write_shapefile_gcs(df, col_map, output_path):
    """Write point shapefile in GCS WGS84 (EPSG:4326)."""
    name_col = col_map["name"]
    has_elev = "elevation" in col_map

    with shapefile.Writer(output_path, shapeType=shapefile.POINT) as w:
        # Define fields
        w.field("Name", "C", size=50)
        w.field("Easting", "N", decimal=3)
        w.field("Northing", "N", decimal=3)
        if has_elev:
            w.field("Elevation", "N", decimal=3)
        w.field("Latitude", "N", decimal=8)
        w.field("Longitude", "N", decimal=8)
        w.field("DMS_Lat", "C", size=30)
        w.field("DMS_Lon", "C", size=30)
        w.field("Remarks", "C", size=30)

        # Add any extra original columns
        extra_cols = [c for c in df.columns
                      if c not in [col_map["name"], col_map["easting"], col_map["northing"],
                                   col_map.get("elevation", ""), "Latitude", "Longitude",
                                   "DMS_Latitude", "DMS_Longitude", "Remarks"]]
        for ec in extra_cols:
            w.field(ec, "C", size=100)

        for _, row in df.iterrows():
            lon = float(row["Longitude"])
            lat = float(row["Latitude"])
            w.point(lon, lat)

            rec = [
                str(row[name_col]),
                float(row[col_map["easting"]]),
                float(row[col_map["northing"]]),
            ]
            if has_elev:
                rec.append(float(row[col_map["elevation"]]))
            rec += [
                float(row["Latitude"]),
                float(row["Longitude"]),
                str(row["DMS_Latitude"]),
                str(row["DMS_Longitude"]),
                str(row["Remarks"]),
            ]
            for ec in extra_cols:
                rec.append(str(row[ec]))

            w.record(*rec)

    # Write .prj file for WGS84
    prj_content = (
        'GEOGCS["GCS_WGS_1984",'
        'DATUM["D_WGS_1984",'
        'SPHEROID["WGS_1984",6378137.0,298.257223563]],'
        'PRIMEM["Greenwich",0.0],'
        'UNIT["Degree",0.0174532925199433]]'
    )
    with open(output_path + ".prj", "w") as f:
        f.write(prj_content)

    print(f"  ✓ GCS Shapefile written: {output_path}.shp")


# ─────────────────────────────────────────────
# 6. WRITE SHAPEFILE (PCS - UTM)
# ─────────────────────────────────────────────

def get_utm_prj(utm_epsg):
    """Return a WKT PRJ string for common UTM WGS84 zones."""
    zone_map = {
        32643: ('WGS_1984_UTM_Zone_43N', 43, 'N'),
        32644: ('WGS_1984_UTM_Zone_44N', 44, 'N'),
        32645: ('WGS_1984_UTM_Zone_45N', 45, 'N'),
        32646: ('WGS_1984_UTM_Zone_46N', 46, 'N'),
        32642: ('WGS_1984_UTM_Zone_42N', 42, 'N'),
        32647: ('WGS_1984_UTM_Zone_47N', 47, 'N'),
        32743: ('WGS_1984_UTM_Zone_43S', 43, 'S'),
        32744: ('WGS_1984_UTM_Zone_44S', 44, 'S'),
    }

    if utm_epsg in zone_map:
        name, zone, hemi = zone_map[utm_epsg]
        cm = (zone * 6) - 183  # central meridian
        return (
            f'PROJCS["{name}",'
            f'GEOGCS["GCS_WGS_1984",'
            f'DATUM["D_WGS_1984",'
            f'SPHEROID["WGS_1984",6378137.0,298.257223563]],'
            f'PRIMEM["Greenwich",0.0],'
            f'UNIT["Degree",0.0174532925199433]],'
            f'PROJECTION["Transverse_Mercator"],'
            f'PARAMETER["False_Easting",500000.0],'
            f'PARAMETER["False_Northing",{"10000000.0" if hemi == "S" else "0.0"}],'
            f'PARAMETER["Central_Meridian",{cm}.0],'
            f'PARAMETER["Scale_Factor",0.9996],'
            f'PARAMETER["Latitude_Of_Origin",0.0],'
            f'UNIT["Meter",1.0]]'
        )
    else:
        # Generic fallback — user should replace with correct PRJ
        return f'PROJCS["EPSG_{utm_epsg}",GEOGCS["GCS_WGS_1984"]]'


def write_shapefile_pcs(df, col_map, utm_epsg, output_path):
    """Write point shapefile in PCS (UTM projection)."""
    name_col = col_map["name"]
    has_elev = "elevation" in col_map

    with shapefile.Writer(output_path, shapeType=shapefile.POINT) as w:
        w.field("Name", "C", size=50)
        w.field("Easting", "N", decimal=3)
        w.field("Northing", "N", decimal=3)
        if has_elev:
            w.field("Elevation", "N", decimal=3)
        w.field("Latitude", "N", decimal=8)
        w.field("Longitude", "N", decimal=8)
        w.field("DMS_Lat", "C", size=30)
        w.field("DMS_Lon", "C", size=30)
        w.field("Remarks", "C", size=30)

        extra_cols = [c for c in df.columns
                      if c not in [col_map["name"], col_map["easting"], col_map["northing"],
                                   col_map.get("elevation", ""), "Latitude", "Longitude",
                                   "DMS_Latitude", "DMS_Longitude", "Remarks"]]
        for ec in extra_cols:
            w.field(ec, "C", size=100)

        for _, row in df.iterrows():
            easting = float(row[col_map["easting"]])
            northing = float(row[col_map["northing"]])
            w.point(easting, northing)

            rec = [
                str(row[name_col]),
                easting,
                northing,
            ]
            if has_elev:
                rec.append(float(row[col_map["elevation"]]))
            rec += [
                float(row["Latitude"]),
                float(row["Longitude"]),
                str(row["DMS_Latitude"]),
                str(row["DMS_Longitude"]),
                str(row["Remarks"]),
            ]
            for ec in extra_cols:
                rec.append(str(row[ec]))

            w.record(*rec)

    # Write .prj file
    with open(output_path + ".prj", "w") as f:
        f.write(get_utm_prj(utm_epsg))

    print(f"  ✓ PCS Shapefile written: {output_path}.shp")


# ─────────────────────────────────────────────
# 7. WRITE EXCEL OUTPUT
# ─────────────────────────────────────────────

def write_excel(df, col_map, output_path):
    """Write full Excel output with formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    def create_sheet(wb, df_subset, sheet_name):
        ws = wb.create_sheet(title=sheet_name)

        # Header style
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Rename columns for display
        display_cols = []
        for c in df_subset.columns:
            if c == col_map["easting"]:
                display_cols.append("Easting")
            elif c == col_map["northing"]:
                display_cols.append("Northing")
            elif c == col_map.get("elevation", ""):
                display_cols.append("Elevation")
            elif c == col_map["name"]:
                display_cols.append("GCP Name")
            elif c == "DMS_Latitude":
                display_cols.append("DMS Latitude")
            elif c == "DMS_Longitude":
                display_cols.append("DMS Longitude")
            else:
                display_cols.append(c)

        # Write header
        for ci, col_name in enumerate(display_cols, start=1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        # Write data
        alt_fill = PatternFill("solid", fgColor="DCE6F1")
        for ri, (_, row) in enumerate(df_subset.iterrows(), start=2):
            fill = alt_fill if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for ci, col_name in enumerate(df_subset.columns, start=1):
                val = row[col_name]
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        # Auto-fit columns
        for ci, col_name in enumerate(display_cols, start=1):
            max_len = max(len(str(col_name)), 10)
            for ri in range(2, ws.max_row + 1):
                val = ws.cell(row=ri, column=ci).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 40)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 20
        return ws

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Sheet 1: All GCPs
    create_sheet(wb, df, "All GCPs")

    # Sheet 2: Temporary GCPs only
    temp_df = df[df["Remarks"] == "Temporary GCP"].reset_index(drop=True)
    if not temp_df.empty:
        create_sheet(wb, temp_df, "Temporary GCPs")

    # Sheet 3: Permanent GCPs only
    perm_df = df[df["Remarks"] == "Permanent GCP"].reset_index(drop=True)
    if not perm_df.empty:
        create_sheet(wb, perm_df, "Permanent GCPs")

    wb.save(output_path)
    print(f"  ✓ Excel written: {output_path}")


# ─────────────────────────────────────────────
# 8. INTERACTIVE PROMPTS
# ─────────────────────────────────────────────

def prompt_input_file():
    """Prompt user for a valid input file path."""
    while True:
        path = input("  Enter input file path (.xlsx / .xls / .csv): ").strip().strip('"').strip("'")
        if not path:
            print("  ✖  Path cannot be empty. Please try again.")
            continue
        if not os.path.exists(path):
            print(f"  ✖  File not found: '{path}'. Please check the path and try again.")
            continue
        ext = os.path.splitext(path)[1].lower()
        if ext not in [".xlsx", ".xls", ".csv"]:
            print(f"  ✖  Unsupported format '{ext}'. Only .xlsx, .xls, .csv are accepted.")
            continue
        return path


def prompt_utm_epsg():
    """Prompt user for a UTM EPSG code with a handy reference list."""
    print()
    print("  Common UTM Zones (WGS84):")
    print("  ┌─────────────┬───────────┐")
    print("  │  Zone       │   EPSG    │")
    print("  ├─────────────┼───────────┤")
    print("  │  UTM 42N    │   32642   │")
    print("  │  UTM 43N    │   32643   │")
    print("  │  UTM 44N    │   32644   │")
    print("  │  UTM 45N    │   32645   │")
    print("  │  UTM 46N    │   32646   │")
    print("  │  UTM 47N    │   32647   │")
    print("  │  UTM 43S    │   32743   │")
    print("  │  UTM 44S    │   32744   │")
    print("  └─────────────┴───────────┘")
    print("  (Any valid EPSG code is accepted)")
    while True:
        raw = input("  Enter UTM EPSG code: ").strip()
        if not raw.isdigit():
            print("  ✖  Please enter a numeric EPSG code (e.g. 32644).")
            continue
        epsg = int(raw)
        # Quick sanity check via pyproj
        try:
            from pyproj import CRS
            CRS.from_epsg(epsg)
            return epsg
        except Exception:
            print(f"  ✖  EPSG:{epsg} is not a recognised projection. Please try again.")


def prompt_mine_name():
    """Prompt user for mine / project name."""
    while True:
        name = input("  Enter mine / project name (used in output filenames): ").strip()
        if not name:
            print("  ✖  Name cannot be empty.")
            continue
        # Replace spaces with underscores for safe filenames
        safe = name.replace(" ", "_")
        if safe != name:
            print(f"  ℹ  Spaces replaced with underscores → '{safe}'")
        return safe


def prompt_output_dir():
    """Prompt user for output directory (default = same folder as script)."""
    raw = input("  Enter output directory [press Enter for current directory]: ").strip().strip('"').strip("'")
    if not raw:
        return "."
    os.makedirs(raw, exist_ok=True)
    return raw


# ─────────────────────────────────────────────
# 9. MAIN
# ─────────────────────────────────────────────

def main():
    print()
    print("╔══════════════════════════════════════════════════════════╗")
    print("║              GCP / PGCP Processor  v1.1                 ║")
    print("║  Outputs: Shapefile (GCS + PCS)  |  Excel               ║")
    print("╚══════════════════════════════════════════════════════════╝")
    print()

    print("── Step 1 of 4 : Input File ─────────────────────────────")
    input_path = prompt_input_file()

    print()
    print("── Step 2 of 4 : UTM Projection ─────────────────────────")
    utm_epsg = prompt_utm_epsg()

    print()
    print("── Step 3 of 4 : Mine / Project Name ────────────────────")
    mine_name = prompt_mine_name()

    print()
    print("── Step 4 of 4 : Output Directory ───────────────────────")
    print("  Two sub-folders will be created inside your chosen directory:")
    print("    GCPs/            -> All GCPs (Temporary + Permanent)")
    print("    Permanent_GCPs/  -> Permanent GCPs only")
    output_dir = prompt_output_dir()

    # Sub-folder paths
    all_dir  = os.path.join(output_dir, "GCPs")
    perm_dir = os.path.join(output_dir, "Permanent_GCPs")
    os.makedirs(all_dir,  exist_ok=True)
    os.makedirs(perm_dir, exist_ok=True)

    # ── Summary before processing ──
    print()
    print("─" * 60)
    print("  Confirm details:")
    print(f"  Input File       : {input_path}")
    print(f"  UTM EPSG         : {utm_epsg}")
    print(f"  Mine Name        : {mine_name}")
    print(f"  All GCPs folder  : {os.path.abspath(all_dir)}")
    print(f"  Perm GCPs folder : {os.path.abspath(perm_dir)}")
    print("─" * 60)
    confirm = input("  Proceed? [Y/n]: ").strip().lower()
    if confirm and confirm not in ("y", "yes"):
        print("  Cancelled.")
        sys.exit(0)

    # ── Read ──
    print()
    print("Reading input file...")
    df, col_map = read_input(input_path)
    print(f"  ✓ {len(df)} records read  |  Columns: {list(df.columns)}")

    # ── Process ──
    print()
    print("Processing coordinates...")
    df = process_data(df, col_map, utm_epsg)
    temp_count = (df["Remarks"] == "Temporary GCP").sum()
    perm_count = (df["Remarks"] == "Permanent GCP").sum()
    print(f"  ✓ Temporary GCPs : {temp_count}")
    print(f"  ✓ Permanent GCPs : {perm_count}")

    # Permanent GCPs subset
    perm_df = df[df["Remarks"] == "Permanent GCP"].reset_index(drop=True)

    # ── Write ALL GCPs outputs ──
    print()
    print("Writing All GCPs outputs  ->  GCPs/")
    write_shapefile_gcs(df, col_map, os.path.join(all_dir, f"3a_{mine_name}_GCP_GCS_MCDR2026"))
    write_shapefile_pcs(df, col_map, utm_epsg, os.path.join(all_dir, f"3b_{mine_name}_GCP_PCS_MCDR2026"))
    write_excel(df, col_map, os.path.join(all_dir, f"3c_{mine_name}_GCP_Details_MCDR2026.xlsx"))

    # ── Write PERMANENT GCPs outputs ──
    print()
    print("Writing Permanent GCPs outputs  ->  Permanent_GCPs/")
    if perm_df.empty:
        print("  ⚠  No Permanent GCPs found — skipping Permanent_GCPs folder.")
    else:
        write_shapefile_gcs(perm_df, col_map, os.path.join(perm_dir, f"3d_{mine_name}_Permanent_GCP_GCS_MCDR2026"))
        write_shapefile_pcs(perm_df, col_map, utm_epsg, os.path.join(perm_dir, f"3e_{mine_name}_Permanent_GCP_PCS_MCDR2026"))
        write_excel(perm_df, col_map, os.path.join(perm_dir, f"3f_{mine_name}_Permanent_GCP_Details_MCDR2026.xlsx"))

    # ── Final summary ──
    print()
    print("╔══════════════════════════════════════════════════════════╗")
    print("║  ✓ Done!  Outputs saved:                                ║")
    print(f"║    GCPs/           -> {temp_count + perm_count} points (All GCPs)            ║")
    print(f"║    Permanent_GCPs/ -> {perm_count} points (Permanent only)       ║")
    print(f"║  Root: {os.path.abspath(output_dir):<52}║")
    print("╚══════════════════════════════════════════════════════════╝")
    print()


if __name__ == "__main__":
    main()
